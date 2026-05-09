import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.price_parser import get_daily_top_pairs


def apply_cell_style(cell, fill=None, font=None, alignment=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def _default_border():
    return Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))


def _default_center():
    return Alignment(horizontal='center', vertical='center')


def _header_fill():
    return PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def _header_font():
    return Font(color="FFFFFF", bold=True)


def generate_price_excel(price_rows, price_headers, times, output_path, discharge_price_ratio):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "电价透视表"

    border = _default_border()
    center_alignment = _default_center()
    header_fill = _header_fill()
    header_font = _header_font()

    num_time_cols = len(times)
    time_col_indices = list(range(num_time_cols))

    col_max = {}
    for col_idx in time_col_indices:
        max_val = 0.0
        for row in price_rows:
            val = row[1 + col_idx] if len(row) > 1 + col_idx else 0.0
            if isinstance(val, (int, float)):
                max_val = max(max_val, val)
        col_max[col_idx] = max_val if max_val > 0 else 1.0

    extended_headers = list(price_headers)
    for i in range(1, 13):
        extended_headers.extend([f"电价差{i}", f"电价差{i}充时", f"电价差{i}放时"])

    for col, header in enumerate(extended_headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_cell_style(cell, fill=header_fill, font=header_font, alignment=center_alignment, border=border)

    for row_idx, row in enumerate(price_rows, start=2):
        prices = row[1:1 + num_time_cols]
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx > 1 and col_idx <= num_time_cols + 1:
                time_idx = col_idx - 2
                max_val = col_max.get(time_idx, 1.0)
                if isinstance(val, (int, float)) and max_val > 0:
                    ratio = min(1.0, val / max_val)
                    r = int(173 - ratio * 50)
                    g = int(216 - ratio * 50)
                    b = 230
                    color = f"{r:02X}{g:02X}{b:02X}"
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.number_format = '0.0000'
            elif col_idx == 1:
                cell.number_format = '@'

        top_pairs = get_daily_top_pairs(prices, discharge_price_ratio, max_pairs=12)
        base_col = num_time_cols + 2
        for k, (spread, ch, dis) in enumerate(top_pairs):
            col_offset = k * 3
            cell_val = ws.cell(row=row_idx, column=base_col + col_offset, value=spread)
            cell_val.number_format = '0.0000'
            cell_val.border = border
            cell_ch = ws.cell(row=row_idx, column=base_col + col_offset + 1, value=times[ch])
            cell_ch.border = border
            cell_ch.number_format = '@'
            cell_dis = ws.cell(row=row_idx, column=base_col + col_offset + 2, value=times[dis])
            cell_dis.border = border
            cell_dis.number_format = '@'

        for k in range(len(top_pairs), 12):
            col_offset = k * 3
            for off in range(3):
                cell = ws.cell(row=row_idx, column=base_col + col_offset + off, value="")
                cell.border = border

    for col in range(1, len(extended_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12 if col <= num_time_cols + 1 else 14

    wb.save(output_path)


def generate_power_matrix_excel(power_matrix_rows, power_matrix_headers, times, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "充放电功率矩阵"

    header_fill = _header_fill()
    header_font = _header_font()
    charge_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    discharge_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    unprofitable_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    unprofitable_font = Font(color="9C0006")
    border = _default_border()
    center_alignment = _default_center()

    for col, header in enumerate(power_matrix_headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_cell_style(cell, fill=header_fill, font=header_font, alignment=center_alignment, border=border)

    num_time_cols = len(times)

    try:
        profit_idx = power_matrix_headers.index("收益(€)") + 1
        threshold_idx = power_matrix_headers.index("折旧及损耗阈值(€/kWh)") + 1
        net_profit_idx = power_matrix_headers.index("实际度电净利(€/kWh)") + 1
    except ValueError:
        profit_idx, threshold_idx, net_profit_idx = -1, -1, -1

    for row_idx, row in enumerate(power_matrix_rows, start=2):
        is_unprofitable = False
        if profit_idx != -1 and isinstance(row[profit_idx - 1], (int, float)):
            if row[profit_idx - 1] <= 0.001:
                is_unprofitable = True

        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx == 1:
                cell.number_format = '@'
            elif col_idx <= num_time_cols + 1:
                if isinstance(val, (int, float)):
                    if val < -0.01:
                        cell.fill = charge_fill
                    elif val > 0.01:
                        cell.fill = discharge_fill
                    cell.number_format = '0'
                else:
                    cell.number_format = '@'
            else:
                if isinstance(val, float):
                    cell.number_format = '0.0000'
                else:
                    cell.number_format = '@'

            if is_unprofitable and col_idx in [profit_idx, threshold_idx, net_profit_idx, 1]:
                cell.fill = unprofitable_fill
                cell.font = unprofitable_font

    for col in range(1, len(power_matrix_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)


def generate_payback_excel(payback_data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投资回收期月报"

    header_fill = _header_fill()
    header_font = _header_font()
    profit_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    total_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    cashflow_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    border = _default_border()
    center_alignment = _default_center()

    year_months = payback_data["sorted_year_months"]
    models = payback_data["models"]
    model_finance = payback_data["model_finance"]

    current_row = 1

    for idx, model in enumerate(models):
        title_cell = ws.cell(row=current_row, column=1, value=f"{model['name']} 投资回收期表")
        title_cell.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=1 + len(year_months))
        current_row += 1

        headers = ["指标"] + year_months
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            apply_cell_style(cell, fill=header_fill, font=header_font, alignment=center_alignment, border=border)
        current_row += 1

        for row_data in model["rows"]:
            label = row_data["label"]
            values = row_data["values"]
            ws.cell(row=current_row, column=1, value=label).border = border
            for col_idx, val in enumerate(values, start=2):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = border
                cell.number_format = '0.00' if isinstance(val, float) else '@'
                if "收益（万欧元）" in label:
                    cell.fill = profit_fill
            current_row += 1

        total_label = model["total_row"]["label"]
        total_value = model["total_row"]["value"]
        ws.cell(row=current_row, column=1, value=total_label).border = border
        total_cell = ws.cell(row=current_row, column=2, value=total_value)
        total_cell.border = border
        total_cell.number_format = '0.00'
        total_cell.fill = total_fill
        for col in range(3, 2 + len(year_months)):
            ws.cell(row=current_row, column=col).border = border
        current_row += 1

        current_row += 1

        finance_title = ws.cell(row=current_row, column=1, value=f"{model['name']} 财务汇总")
        finance_title.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=1 + len(year_months))
        current_row += 1

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            apply_cell_style(cell, fill=header_fill, font=header_font, alignment=center_alignment, border=border)
        current_row += 1

        finance_info = model_finance[model["name"]]
        for label, values in finance_info["finance_rows"]:
            ws.cell(row=current_row, column=1, value=label).border = border
            for col_idx, val in enumerate(values, start=2):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = border
                cell.number_format = '0.00' if isinstance(val, float) else '@'
            current_row += 1

        cum_label, cum_values = finance_info["cum_cashflow_row"]
        ws.cell(row=current_row, column=1, value=cum_label).border = border
        for col_idx, val in enumerate(cum_values, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = border
            cell.number_format = '0.00'
            cell.fill = cashflow_fill
        current_row += 1

        payback_value = finance_info["payback"]
        payback_label = f"{model['name']} 回收周期（月）"
        ws.cell(row=current_row, column=1, value=payback_label).border = border
        payback_cell = ws.cell(row=current_row, column=2,
                               value=payback_value if payback_value is not None else "未回本")
        payback_cell.border = border
        if payback_value is not None:
            payback_cell.number_format = '0.00'
        payback_cell.fill = cashflow_fill
        for col in range(3, 2 + len(year_months)):
            ws.cell(row=current_row, column=col).border = border
        current_row += 1

        if idx < len(models) - 1:
            current_row += 2

    for col in range(1, 2 + len(year_months)):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output_path)
